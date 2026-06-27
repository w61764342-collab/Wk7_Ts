#!/usr/bin/env python3
"""
Validate KCSB Excel/PDF files in Cloudflare R2 against websites-config.yml schemas.

Reports and stats are written to R2 only — never to the local repo.
"""

from __future__ import annotations

import argparse
import fnmatch
import io
import json
import logging
import os
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import PurePosixPath
from typing import Any

import boto3
import yaml
from botocore.config import Config
from botocore.exceptions import ClientError
import xlrd
from openpyxl import load_workbook

MONITOR_PREFIX = "KCSB-Data/monitor"
CONFIG_R2_KEY = f"{MONITOR_PREFIX}/websites-config.yml"
SEVERITY_CRITICAL = "critical"
SEVERITY_HIGH = "high"
SEVERITY_MEDIUM = "medium"

logger = logging.getLogger(__name__)


def setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s - %(levelname)s - %(message)s",
        stream=sys.stdout,
        force=True,
    )


def load_config(client: Any, bucket: str, local_path: str | None) -> dict:
    """Load websites-config.yml from R2 (default) or a local file (--config-local)."""
    if local_path:
        with open(local_path, encoding="utf-8") as fh:
            data = yaml.safe_load(fh)
        print(f"Config loaded from local file: {local_path}")
        logger.info("Config loaded from local file: %s", local_path)
        return data

    try:
        body = download_object(client, bucket, CONFIG_R2_KEY)
    except ClientError as exc:
        if exc.response["Error"]["Code"] in ("404", "NoSuchKey"):
            raise SystemExit(
                f"websites-config.yml not found at s3://{bucket}/{CONFIG_R2_KEY}. "
                "Upload the config to R2 under KCSB-Data/monitor/, "
                "or pass --config-local for a local file."
            ) from exc
        raise

    logger.info("Downloading config from s3://%s/%s", bucket, CONFIG_R2_KEY)
    print(f"Config loaded from s3://{bucket}/{CONFIG_R2_KEY}")
    return yaml.safe_load(body)


def build_r2_client() -> Any:
    required = [
        "CF_R2_ACCESS_KEY_ID",
        "CF_R2_SECRET_ACCESS_KEY",
        "CF_R2_ENDPOINT_URL",
        "CF_R2_BUCKET_NAME",
    ]
    missing = [v for v in required if not os.environ.get(v)]
    if missing:
        raise SystemExit(f"Missing environment variables: {', '.join(missing)}")

    return boto3.client(
        "s3",
        endpoint_url=os.environ["CF_R2_ENDPOINT_URL"],
        aws_access_key_id=os.environ["CF_R2_ACCESS_KEY_ID"],
        aws_secret_access_key=os.environ["CF_R2_SECRET_ACCESS_KEY"],
        region_name="us-east-1",
        config=Config(
            signature_version="s3v4",
            s3={"addressing_style": "path"},
        ),
    )


def normalise_prefix(prefix: str) -> str:
    prefix = prefix.lstrip("/")
    if prefix and not prefix.endswith("/"):
        prefix += "/"
    return prefix


def list_objects(client: Any, bucket: str, prefix: str) -> list[dict]:
    prefix = normalise_prefix(prefix)
    paginator = client.get_paginator("list_objects_v2")
    objects: list[dict] = []
    page_num = 0
    for page in paginator.paginate(Bucket=bucket, Prefix=prefix):
        page_num += 1
        batch = page.get("Contents", [])
        objects.extend(batch)
        logger.info(
            "  Listed page %d: %d objects (total so far: %d, truncated=%s)",
            page_num,
            len(batch),
            len(objects),
            page.get("IsTruncated", False),
        )
    return objects


def download_object(client: Any, bucket: str, key: str) -> bytes:
    response = client.get_object(Bucket=bucket, Key=key)
    return response["Body"].read()


def load_existing_stats(client: Any, bucket: str) -> dict:
    key = f"{MONITOR_PREFIX}/monitor_stats.yml"
    try:
        logger.info("Loading existing stats from s3://%s/%s", bucket, key)
        body = download_object(client, bucket, key)
        return yaml.safe_load(body) or {}
    except ClientError as exc:
        if exc.response["Error"]["Code"] in ("404", "NoSuchKey"):
            logger.info("No existing stats file — starting fresh")
            return {}
        raise


def log_failed_checks(file_label: str, checks: list[dict]) -> None:
    for check in checks:
        if not check.get("passed"):
            logger.warning(
                "  FAIL %s: %s — %s",
                file_label,
                check.get("check"),
                check.get("detail"),
            )


def upload_text(client: Any, bucket: str, key: str, body: str, content_type: str) -> None:
    client.put_object(
        Bucket=bucket,
        Key=key,
        Body=body.encode("utf-8"),
        ContentType=content_type,
    )


def tab_from_key(key: str) -> str:
    parts = PurePosixPath(key).parts
    if len(parts) >= 2:
        return parts[-2]
    return ""


def match_profile(key: str, profiles: list[dict]) -> dict | None:
    filename = PurePosixPath(key).name
    tab = tab_from_key(key)

    for profile in profiles:
        pattern = profile.get("excel_file_pattern", "*.xlsx")
        exclude = profile.get("exclude_pattern")
        if exclude and fnmatch.fnmatch(filename, exclude):
            continue
        if not fnmatch.fnmatch(filename, pattern):
            continue
        match_tab = profile.get("match_tab")
        if match_tab and tab != match_tab:
            continue
        return profile
    return None


@dataclass
class SheetView:
    name: str
    headers: list[str]
    rows: list[tuple[Any, ...]]
    data_row_count: int


def detect_spreadsheet_format(content: bytes) -> str:
    """Return xlsx, xls, pdf, html, or unknown based on file magic bytes."""
    if len(content) < 4:
        return "unknown"
    if content[:2] == b"PK":
        return "xlsx"
    if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    if content[:4] == b"%PDF":
        return "pdf"
    head = content[:256].lstrip().lower()
    if head.startswith(b"<html") or head.startswith(b"<!doctype"):
        return "html"
    return "unknown"


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_xlsx_sheet(wb: Any, sheet_name: str) -> SheetView:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = [_cell_str(c) for c in (rows[0] if rows else [])]
    return SheetView(
        name=sheet_name,
        headers=headers,
        rows=rows,
        data_row_count=max(len(rows) - 1, 0),
    )


def read_xls_sheet(book: xlrd.book.Book, sheet_name: str) -> SheetView:
    sheet = book.sheet_by_name(sheet_name)
    rows = [
        tuple(_cell_str(sheet.cell_value(r, c)) for c in range(sheet.ncols))
        for r in range(sheet.nrows)
    ]
    headers = list(rows[0]) if rows else []
    return SheetView(
        name=sheet_name,
        headers=headers,
        rows=rows,
        data_row_count=max(len(rows) - 1, 0),
    )


def open_spreadsheet(content: bytes) -> tuple[str, list[str], Any]:
    """
    Open workbook bytes. Returns (format, sheet_names, handle).
    handle is openpyxl Workbook or xlrd Book — caller must close/release.
    """
    fmt = detect_spreadsheet_format(content)
    if fmt == "xlsx":
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        return fmt, wb.sheetnames, wb
    if fmt == "xls":
        book = xlrd.open_workbook(file_contents=content, on_demand=True)
        return fmt, book.sheet_names(), book
    raise ValueError(f"unsupported spreadsheet format: {fmt}")


def close_spreadsheet(fmt: str, handle: Any) -> None:
    if fmt == "xlsx":
        handle.close()
    elif fmt == "xls":
        handle.release_resources()


def get_sheet_view(fmt: str, handle: Any, sheet_name: str) -> SheetView:
    if fmt == "xlsx":
        return read_xlsx_sheet(handle, sheet_name)
    return read_xls_sheet(handle, sheet_name)


def inspect_excel(
    content: bytes,
    profile: dict,
    quality: bool,
) -> tuple[list[dict], bool]:
    checks: list[dict] = []
    all_passed = True

    min_kb = profile.get("min_file_size_kb", 5)
    size_kb = len(content) / 1024
    size_ok = size_kb >= min_kb
    checks.append({
        "check": "min_file_size_kb",
        "severity": SEVERITY_HIGH,
        "passed": size_ok,
        "detail": f"{size_kb:.1f} KB (min {min_kb} KB)",
    })
    all_passed &= size_ok

    detected = detect_spreadsheet_format(content)
    if detected in ("pdf", "html", "unknown"):
        detail = {
            "pdf": "file is PDF, not Excel",
            "html": "file is HTML (likely a failed download page)",
            "unknown": "unrecognised file format",
        }[detected]
        checks.append({
            "check": "file_readable",
            "severity": SEVERITY_CRITICAL,
            "passed": False,
            "detail": detail,
        })
        return checks, False

    try:
        fmt, sheet_names, handle = open_spreadsheet(content)
    except Exception as exc:
        checks.append({
            "check": "file_readable",
            "severity": SEVERITY_CRITICAL,
            "passed": False,
            "detail": str(exc),
        })
        return checks, False

    reader = "xlrd (.xls)" if fmt == "xls" else "openpyxl (.xlsx)"
    format_detail = (
        f"legacy Excel 97-2003 (.xls) opened with {reader}"
        if fmt == "xls"
        else f"Office Open XML (.xlsx) opened with {reader}"
    )
    checks.append({
        "check": "file_format",
        "severity": SEVERITY_MEDIUM,
        "passed": True,
        "detail": format_detail,
    })
    checks.append({
        "check": "file_readable",
        "severity": SEVERITY_CRITICAL,
        "passed": True,
        "detail": format_detail,
    })

    if not sheet_names:
        checks.append({
            "check": "has_sheets",
            "severity": SEVERITY_CRITICAL,
            "passed": False,
            "detail": "workbook has no sheets",
        })
        close_spreadsheet(fmt, handle)
        return checks, False

    sheet_specs = profile.get("sheets", [])
    try:
        for spec in sheet_specs:
            expected_name = spec.get("name", "*")
            required_cols = spec.get("required_columns") or []
            row_min, row_max = spec.get("row_count_range", [1, 999999])

            if expected_name == "*":
                target_sheet = sheet_names[0]
            elif expected_name not in sheet_names:
                checks.append({
                    "check": f"sheet_exists:{expected_name}",
                    "severity": SEVERITY_CRITICAL,
                    "passed": False,
                    "detail": f"expected '{expected_name}', found {sheet_names}",
                })
                all_passed = False
                continue
            else:
                target_sheet = expected_name

            view = get_sheet_view(fmt, handle, target_sheet)
            headers = view.headers
            data_rows = view.data_row_count
            rows = view.rows

            if required_cols:
                missing = [c for c in required_cols if c not in headers]
                cols_ok = not missing
                checks.append({
                    "check": f"required_columns:{target_sheet}",
                    "severity": SEVERITY_CRITICAL,
                    "passed": cols_ok,
                    "detail": (
                        f"missing {missing}"
                        if missing
                        else f"all {len(required_cols)} columns present"
                    ),
                })
                all_passed &= cols_ok

            rows_ok = row_min <= data_rows <= row_max
            checks.append({
                "check": f"row_count_range:{target_sheet}",
                "severity": SEVERITY_MEDIUM,
                "passed": rows_ok,
                "detail": f"{data_rows} rows (expected {row_min}–{row_max})",
            })
            all_passed &= rows_ok

            if quality and required_cols and data_rows > 0:
                col_idx = {h: i for i, h in enumerate(headers)}
                for col in required_cols:
                    if col not in col_idx:
                        continue
                    idx = col_idx[col]
                    nulls = sum(
                        1
                        for row in rows[1:]
                        if idx >= len(row) or row[idx] is None or _cell_str(row[idx]) == ""
                    )
                    null_pct = (nulls / data_rows) * 100
                    null_ok = null_pct <= 50
                    checks.append({
                        "check": f"null_pct:{col}",
                        "severity": SEVERITY_MEDIUM,
                        "passed": null_ok,
                        "detail": f"{null_pct:.1f}% null/empty in '{col}'",
                    })
                    all_passed &= null_ok
    finally:
        close_spreadsheet(fmt, handle)

    return checks, all_passed


def inspect_pdf(content: bytes, pdf_schema: dict) -> tuple[list[dict], bool]:
    checks: list[dict] = []
    all_passed = True

    min_kb = pdf_schema.get("min_file_size_kb", 10)
    size_kb = len(content) / 1024
    size_ok = size_kb >= min_kb
    checks.append({
        "check": "min_file_size_kb",
        "severity": SEVERITY_HIGH,
        "passed": size_ok,
        "detail": f"{size_kb:.1f} KB (min {min_kb} KB)",
    })
    all_passed &= size_ok

    magic_ok = content[:4] == b"%PDF"
    checks.append({
        "check": "pdf_magic_bytes",
        "severity": SEVERITY_CRITICAL,
        "passed": magic_ok,
        "detail": "starts with %PDF" if magic_ok else "invalid PDF header",
    })
    all_passed &= magic_ok

    return checks, all_passed


def merge_stats(existing: dict, scraper: str, observations: dict) -> dict:
    stats = dict(existing)
    scraper_stats = dict(stats.get(scraper, {}))
    for key, value in observations.items():
        if key.endswith("_min"):
            scraper_stats[key] = min(scraper_stats.get(key, value), value)
        elif key.endswith("_max"):
            scraper_stats[key] = max(scraper_stats.get(key, value), value)
        elif key == "columns_seen":
            seen = set(scraper_stats.get("columns_seen", []))
            seen.update(value)
            scraper_stats["columns_seen"] = sorted(seen)
        else:
            scraper_stats[key] = value
    stats[scraper] = scraper_stats
    return stats


def write_step_summary(lines: list[str]) -> None:
    summary_path = os.environ.get("GITHUB_STEP_SUMMARY")
    if not summary_path:
        return
    with open(summary_path, "a", encoding="utf-8") as fh:
        fh.write("\n".join(lines) + "\n")


def run_validation(args: argparse.Namespace) -> int:
    logger.info("=" * 60)
    logger.info("KCSB R2 Schema Monitor — starting")
    logger.info("Bucket: %s", os.environ.get("CF_R2_BUCKET_NAME", "(not set)"))
    logger.info(
        "Options: update_stats=%s quality=%s fail_on_error=%s category=%s date=%s",
        args.update_stats,
        args.quality,
        args.fail_on_error,
        args.category or "all",
        args.date or "today UTC",
    )
    logger.info("=" * 60)

    client = build_r2_client()
    bucket = os.environ["CF_R2_BUCKET_NAME"]
    config = load_config(client, bucket, args.config_local)

    scrapers = {s["name"]: s for s in config.get("scrapers", [])}
    schemas = {e["scraper"]: e for e in config.get("excel_schema", [])}
    logger.info("Config: %d scrapers, %d excel_schema entries", len(scrapers), len(schemas))

    if not scrapers:
        logger.error("No scrapers defined in config — nothing to validate")
        return 1

    report_date = args.date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    report: dict[str, Any] = {
        "report_date": report_date,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "schedule": "quarterly",
        "scrapers": {},
    }

    existing_stats = load_existing_stats(client, bucket) if args.update_stats else {}
    updated_stats = dict(existing_stats)
    if args.update_stats:
        logger.info("Loaded stats for %d scraper(s)", len(existing_stats))

    summary_rows: list[str] = [
        "## KCSB R2 Schema Monitor",
        "",
        f"**Report date:** {report_date}",
        "",
        "| Scraper | XLSX | PDF | Checks | Status |",
        "|---------|------|-----|--------|--------|",
    ]

    any_failure = False

    target_scrapers = list(scrapers.keys())
    if args.category:
        matched = [n for n, s in scrapers.items() if s.get("display_name") == args.category]
        if not matched:
            print(f"ERROR: unknown category '{args.category}'", file=sys.stderr)
            return 1
        target_scrapers = matched

    logger.info("Validating %d scraper(s): %s", len(target_scrapers), ", ".join(target_scrapers))

    for scraper_idx, scraper_name in enumerate(target_scrapers, 1):
        scraper_cfg = scrapers[scraper_name]
        schema_cfg = schemas.get(scraper_name, {})
        prefix = normalise_prefix(scraper_cfg["r2_path"])
        display_name = scraper_cfg.get("display_name", scraper_name)

        logger.info("")
        logger.info(
            "[%d/%d] Scraper: %s (%s)",
            scraper_idx,
            len(target_scrapers),
            display_name,
            scraper_name,
        )
        logger.info("  R2 prefix: s3://%s/%s", bucket, prefix)

        if not schema_cfg:
            logger.warning("  No excel_schema entry for '%s' — file checks will be limited", scraper_name)

        logger.info("  Listing objects in R2...")
        objects = list_objects(client, bucket, prefix)
        xlsx_keys = [o["Key"] for o in objects if o["Key"].lower().endswith(".xlsx")]
        pdf_keys = [o["Key"] for o in objects if o["Key"].lower().endswith(".pdf")]
        logger.info(
            "  Found %d total objects — %d xlsx, %d pdf",
            len(objects),
            len(xlsx_keys),
            len(pdf_keys),
        )

        profiles = schema_cfg.get("profiles", [])
        pdf_schema = schema_cfg.get("pdf_schema", {"min_file_size_kb": 10})
        expectations = schema_cfg.get("expectations", {})

        file_count = len(xlsx_keys) + len(pdf_keys)
        scraper_result: dict[str, Any] = {
            "r2_prefix": prefix,
            "files_found": file_count,
            "xlsx_count": len(xlsx_keys),
            "pdf_count": len(pdf_keys),
            "unique_ads": file_count,
            "total_rows": file_count,
            "ads_source": "file_count",
            "files": [],
            "category_checks": [],
            "checks_passed": 0,
            "checks_total": 0,
            "all_passed": True,
        }

        min_xlsx = expectations.get("min_xlsx_files", 0)
        min_pdf = expectations.get("min_pdf_files", 0)
        for label, count, minimum, severity in [
            ("min_xlsx_files", len(xlsx_keys), min_xlsx, SEVERITY_HIGH),
            ("min_pdf_files", len(pdf_keys), min_pdf, SEVERITY_MEDIUM),
        ]:
            passed = count >= minimum
            scraper_result["category_checks"].append({
                "check": label,
                "severity": severity,
                "passed": passed,
                "detail": f"{count} found (min {minimum})",
            })
            scraper_result["checks_total"] += 1
            if passed:
                scraper_result["checks_passed"] += 1
            else:
                scraper_result["all_passed"] = False
                any_failure = True
            level = logger.info if passed else logger.warning
            level("  Category check %s: %s", label, scraper_result["category_checks"][-1]["detail"])

        observations: dict[str, Any] = {
            "last_run": report_date,
            "xlsx_count": len(xlsx_keys),
            "pdf_count": len(pdf_keys),
        }

        file_total = len(xlsx_keys) + len(pdf_keys)
        file_num = 0

        for key in xlsx_keys:
            file_num += 1
            filename = PurePosixPath(key).name
            logger.info("  [%d/%d] Inspecting xlsx: %s", file_num, file_total, filename)
            logger.debug("    Full key: %s", key)
            content = download_object(client, bucket, key)
            profile = match_profile(key, profiles)
            file_entry: dict[str, Any] = {
                "key": key,
                "type": "xlsx",
                "size_kb": round(len(content) / 1024, 1),
                "profile_matched": bool(profile),
                "checks": [],
                "passed": True,
            }

            if not profile:
                file_entry["checks"].append({
                    "check": "profile_match",
                    "severity": SEVERITY_MEDIUM,
                    "passed": False,
                    "detail": "no matching excel_schema profile",
                })
                file_entry["passed"] = False
                scraper_result["all_passed"] = False
                any_failure = True
                logger.warning("    FAIL — no matching excel_schema profile")
            else:
                checks, passed = inspect_excel(content, profile, args.quality)
                file_entry["checks"] = checks
                file_entry["passed"] = passed
                if not passed:
                    scraper_result["all_passed"] = False
                    any_failure = True
                    log_failed_checks(filename, checks)
                else:
                    logger.info("    PASS — %d check(s)", len(checks))

            scraper_result["checks_total"] += len(file_entry["checks"])
            scraper_result["checks_passed"] += sum(1 for c in file_entry["checks"] if c["passed"])
            scraper_result["files"].append(file_entry)

            observations["file_size_kb_min"] = min(
                observations.get("file_size_kb_min", file_entry["size_kb"]),
                file_entry["size_kb"],
            )
            observations["file_size_kb_max"] = max(
                observations.get("file_size_kb_max", file_entry["size_kb"]),
                file_entry["size_kb"],
            )

        for key in pdf_keys:
            file_num += 1
            filename = PurePosixPath(key).name
            logger.info("  [%d/%d] Inspecting pdf: %s", file_num, file_total, filename)
            logger.debug("    Full key: %s", key)
            content = download_object(client, bucket, key)
            checks, passed = inspect_pdf(content, pdf_schema)
            file_entry = {
                "key": key,
                "type": "pdf",
                "size_kb": round(len(content) / 1024, 1),
                "checks": checks,
                "passed": passed,
            }
            if not passed:
                scraper_result["all_passed"] = False
                any_failure = True
                log_failed_checks(filename, checks)
            else:
                logger.info("    PASS — %d check(s)", len(checks))
            scraper_result["checks_total"] += len(checks)
            scraper_result["checks_passed"] += sum(1 for c in checks if c["passed"])
            scraper_result["files"].append(file_entry)

        report["scrapers"][scraper_name] = scraper_result

        status = "PASS" if scraper_result["all_passed"] else "FAIL"
        logger.info(
            "  Scraper result: %s — %d/%d checks passed (%d xlsx, %d pdf)",
            status,
            scraper_result["checks_passed"],
            scraper_result["checks_total"],
            len(xlsx_keys),
            len(pdf_keys),
        )
        summary_rows.append(
            f"| {scraper_cfg.get('display_name', scraper_name)} "
            f"| {len(xlsx_keys)} | {len(pdf_keys)} "
            f"| {scraper_result['checks_passed']}/{scraper_result['checks_total']} "
            f"| {status} |"
        )

        if args.update_stats:
            updated_stats = merge_stats(updated_stats, scraper_name, observations)

    report["total_unique_ads"] = sum(
        s.get("unique_ads") or 0 for s in report["scrapers"].values()
    )

    report_key = f"{MONITOR_PREFIX}/{report_date}/report.json"
    logger.info("")
    logger.info("Uploading report to s3://%s/%s", bucket, report_key)
    upload_text(client, bucket, report_key, json.dumps(report, ensure_ascii=False, indent=2), "application/json")
    logger.info("Report uploaded successfully")

    if args.update_stats:
        stats_key = f"{MONITOR_PREFIX}/monitor_stats.yml"
        logger.info("Uploading stats to s3://%s/%s", bucket, stats_key)
        upload_text(
            client,
            bucket,
            stats_key,
            yaml.dump(updated_stats, allow_unicode=True, default_flow_style=False),
            "text/yaml",
        )
        logger.info("Stats uploaded successfully")

    logger.info("")
    logger.info("=" * 60)
    logger.info("SUMMARY")
    for row in summary_rows:
        if row.startswith("|") and not row.startswith("|-"):
            logger.info(row.replace("|", " ").strip())
        elif row.startswith("##") or row.startswith("**"):
            logger.info(row.replace("*", "").strip())
    overall = "FAIL" if any_failure else "PASS"
    logger.info("Overall result: %s", overall)
    logger.info("=" * 60)

    for row in summary_rows:
        print(row.replace("|", " ").strip())
    write_step_summary(summary_rows)

    if args.fail_on_error and any_failure:
        return 1
    return 0


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Validate KCSB R2 files against websites-config.yml (loaded from R2 by default)"
    )
    parser.add_argument(
        "--config-local",
        default=None,
        metavar="PATH",
        help=f"Use a local config file instead of R2 (default: s3://…/{CONFIG_R2_KEY})",
    )
    parser.add_argument("--date", default=None, help="Report date label YYYY-MM-DD (default: today UTC)")
    parser.add_argument("--category", default=None, help="Filter by main category display name (Arabic)")
    parser.add_argument("--update-stats", action="store_true", help="Merge observations into monitor_stats.yml in R2")
    parser.add_argument("--quality", action="store_true", help="Run deep null-percentage checks on text Excel files")
    parser.add_argument("--fail-on-error", action="store_true", help="Exit 1 if any check failed")
    parser.add_argument("--verbose", "-v", action="store_true", help="Debug logging (full R2 keys per file)")
    args = parser.parse_args()
    setup_logging(args.verbose)
    raise SystemExit(run_validation(args))


if __name__ == "__main__":
    main()
